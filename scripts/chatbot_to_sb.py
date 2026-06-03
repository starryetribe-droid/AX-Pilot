#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""챗봇 변환기 v2: 보완 시트(08 메타 + 09 렌더계약 + 10 바인딩) → AX Pilot screens.json.

시트09 '렌더아키타입/역할(role)'을 단일 진실로 삼아, chatbot_components 아키타입
디스패치로 KRDS 마크업을 생성한다. 변환기에는 컴포넌트별 마크업 하드코딩이 없다.
데이터(09 계약 + 10 바인딩)만으로 모듈별 UI를 생성하는 변환기."""
import openpyxl, json, re, sys, os
from collections import defaultdict
HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
import chatbot_components as cc

SUP = os.environ.get('CHATBOT_XLSX',
                     '/Users/Starry/Downloads/풀무원_디자인밀_AI챗봇_인텐트정의_보완_20260602.xlsx')
TARGET = int(sys.argv[1]) if len(sys.argv) > 1 else 7
BUILD_DIR = os.path.join(HERE, '_build')
os.makedirs(BUILD_DIR, exist_ok=True)
OUT_JSON = os.path.join(BUILD_DIR, f'IT{TARGET:03d}_screens.json')

# --- 와이어프레임 시각화용 더미 샘플(실제값은 런타임 바인딩). 필드명 suffix 기준 ---
FIELD_SAMPLE = {
    'balance': '1,250 P', 'expiring': '80 P', 'scheduled': '300 P',
    'point': '1,250 P', 'amount': '38,000원',
    'result': 'success', 'status': 'success',
    'maxQuantity': '5', 'currentQty': '2', 'newQty': '3', 'quantity': '2',
    'selectedProduct': '유기농 콩나물 1+1', 'productName': '유기농 콩나물 1+1',
    'name': '유기농 콩나물 1+1', 'title': '유기농 콩나물 1+1',
    'deliveryDate': '6/10(화)', 'orderDate': '6/1(일)', 'date': '6/10(화)',
    'orderId': 'DM240601-0001', 'price': '12,900원',
}
def sample(v):
    s = (v or '').strip()
    def repl(m):
        key = m.group(1).split('.')[-1].rstrip('[]').strip()
        return FIELD_SAMPLE.get(key, '○○')
    return re.sub(r'\{\{\s*([^}]+?)\s*\}\}', repl, s)
h = cc.Helpers(sample)

wb = openpyxl.load_workbook(SUP)
ws08, ws09, ws10 = wb['8_시나리오모듈'], wb['09_컴포넌트스키마'], wb['10_모듈콘텐츠바인딩']

# --- 09: 렌더 계약(단일 진실): 컴포넌트→아키타입, (컴포넌트,슬롯)→역할, 컴포넌트→명칭 ---
ARCH, ROLE, CNAME = {}, {}, {}
for r in range(3, ws09.max_row + 1):
    comp = ws09.cell(r, 1).value
    if not comp:
        continue
    CNAME[comp] = ws09.cell(r, 2).value
    slot, arch, role = ws09.cell(r, 3).value, ws09.cell(r, 10).value, ws09.cell(r, 11).value
    if arch:
        ARCH[comp] = arch
    if slot and role:
        ROLE[(comp, slot)] = role

# --- 08: 대상 인텐트 모듈 메타(단계순) ---
modules = []
for r in range(3, ws08.max_row + 1):
    if ws08.cell(r, 2).value == TARGET:
        modules.append(dict(
            code=ws08.cell(r, 1).value, intent=ws08.cell(r, 2).value,
            intent_name=ws08.cell(r, 3).value, step=ws08.cell(r, 4).value,
            name=ws08.cell(r, 5).value, layer=ws08.cell(r, 6).value,
            atype=ws08.cell(r, 7).value, comps=ws08.cell(r, 8).value,
            cond=ws08.cell(r, 9).value, api=ws08.cell(r, 10).value, exc=ws08.cell(r, 11).value))
modules.sort(key=lambda m: str(m['step']))
assert modules, f'인텐트 #{TARGET} 모듈이 08 시트에 없음'

# --- 10: 모듈별 슬롯 바인딩(행 순서 보존) ---
binds = defaultdict(list)   # code -> [(comp, slot, srctype, value, note)]
for r in range(3, ws10.max_row + 1):
    code = ws10.cell(r, 1).value
    if not code or not str(code).startswith('IT') or ws10.cell(r, 2).value != TARGET:
        continue
    binds[code].append((ws10.cell(r, 5).value, ws10.cell(r, 6).value,
                         ws10.cell(r, 7).value, ws10.cell(r, 8).value, ws10.cell(r, 9).value))

def comp_order(code):
    """렌더 대상 컴포넌트 = 10시트 바인딩 등장 순서(단일 소스). 없으면 8시트 comps 폴백."""
    seen = []
    for (comp, *_rest) in binds[code]:
        if comp and comp not in seen:
            seen.append(comp)
    if not seen:
        meta = next((m for m in modules if m['code'] == code), None)
        if meta and meta['comps']:
            seen = [c.strip() for c in str(meta['comps']).replace('+', ' ').split() if c.strip()]
    return seen

def roles_of(code, comp):
    """컴포넌트의 슬롯 바인딩을 09 계약의 role 키로 매핑(계약에 없으면 슬롯명 그대로)."""
    roles = {}
    for (c, slot, _srctype, value, _note) in binds[code]:
        if c == comp:
            roles[ROLE.get((c, slot), slot)] = value
    return roles

# ---------- desc-panel (AX Pilot 규칙: 한 항목 = 하나의 desc-block) ----------
def db(parts):  return '<li><p class="desc-block">' + '<br>'.join(parts) + '</p></li>'
def lvl1(t):    return f'<span class="lvl1">{cc.html.escape(t)}</span>'
def lvl2(t):    return f'<span class="lvl2">&nbsp;&nbsp;&nbsp;• {cc.html.escape(t)}</span>'

def section(title, body):
    return (f'<div class="desc-section"><div class="desc-section-header">{cc.html.escape(title)}</div>'
            f'<div class="desc-section-body">{body}</div></div>')

def desc_panel(m, ui_items):
    info = section('화면 정보', '<ul class="desc-list">'
                   + db([lvl1(f"{m['code']} · {m['name']}")])
                   + db([lvl2(f"답변유형: {m['atype']}"), lvl2(f"계층: {m['layer']}"), lvl2(f"단계: {m['step']}")])
                   + '</ul>')
    ui_lis = ''
    for n, label, binds_desc in ui_items:
        parts = [lvl1(label)] + [lvl2(bd) for bd in binds_desc]
        ui_lis += db(parts)
    ui = section('UI 요소', f'<ul class="desc-list">{ui_lis}</ul>')
    cond = section('노출 조건', '<ul class="desc-list">' + db([lvl1(str(m['cond'] or '-'))]) + '</ul>')
    api = section('API / 데이터', '<ul class="desc-list">'
                  + db([lvl1(f"호출 API: {m['api']}" if m['api'] else '호출 API: 없음')]) + '</ul>')
    exc = section('예외 처리', '<ul class="desc-list">'
                  + db([lvl1(str(m['exc'] or '정의된 예외 분기 없음'))]) + '</ul>')
    return f'<div class="desc-panel">{info}{ui}{cond}{api}{exc}</div>'

SRC_TAG = {'static': '고정', 'api': 'API', 'session': '세션', 'rag': 'RAG', 'system': '시스템'}

# ---------- 화면 생성 ----------
screens = []
prev_user = str(modules[0]['intent_name'])   # 인텐트 진입 트리거 발화(대표)
for m in modules:
    code = m['code']
    msgs, n, ui_items = '', 1, []
    is_first = str(m['step']) in ('01', '1')
    if is_first:
        msgs += cc.user_row(prev_user, n)
        ui_items.append((n, f'{n}. 사용자 발화 — 인텐트 진입 트리거', ['source: 사용자 입력 / 인텐트 분류']))
        n += 1
    else:
        msgs += cc.user_row(prev_user, None, faded=True)   # 맥락(번호 없음)

    for comp in comp_order(code):
        arch = ARCH.get(comp)
        renderer = cc.ARCHETYPES.get(arch)
        if not renderer:
            continue
        roles = roles_of(code, comp)
        msgs += cc.bot_row(renderer(roles, h), n)
        bds = [f"{slot} ({SRC_TAG.get(srctype, srctype)}): {value}"
               for (c, slot, srctype, value, _note) in binds[code] if c == comp]
        cname = CNAME.get(comp, comp)
        ui_items.append((n, f'{n}. {comp} {cname} — {m["name"]}', bds))
        n += 1

    body = cc.wf_panel(msgs) + desc_panel(m, ui_items)
    screens.append(dict(id=code, title=f'{m["intent_name"]} - {m["name"]}',
                        path=f'AI 챗봇 > {m["intent_name"]}', body=body))

data = dict(feature_id=f'IT{TARGET:03d}', feature_name=str(modules[0]['intent_name']),
            author=os.environ.get('CHATBOT_AUTHOR', 'AX Pilot'), ymd='2026-06', screens=screens)
with open(OUT_JSON, 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=2)
print('intent:', TARGET, modules[0]['intent_name'])
for m in modules:
    print(f"  {m['code']:18s} step={m['step']:>2}  comps={comp_order(m['code'])}  "
          f"arch={[ARCH.get(c) for c in comp_order(m['code'])]}")
print('JSON:', OUT_JSON)
