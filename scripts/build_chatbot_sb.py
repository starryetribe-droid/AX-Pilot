#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""챗봇 SB 원클릭 빌드: 인텐트 번호 1개 → SB HTML 일체 생성.

  python3 chatbot_sb/build_chatbot_sb.py 7

단계: ① chatbot_to_sb.py (워크북 09계약+10바인딩 → screens.json)
      ② generate_sb.py  (screens.json → 화면별 SB HTML)
      ③ index.html      (라벨 단 iframe 묶음)
출력: ~/Downloads/SB_챗봇/IT### (Finder에서 열람) + /tmp/sb_view/IT### (미리보기 서버용 미러)
     · CHATBOT_OUT 환경변수로 출력 루트 변경 가능(기본 ~/Downloads/SB_챗봇)
"""
import os, sys, subprocess, shutil
from collections import defaultdict
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
SKILL = os.path.expanduser('~/.claude/skills/feature-spec')
GEN = os.path.join(SKILL, 'scripts', 'generate_sb.py')
XLSX = os.environ.get('CHATBOT_XLSX',
                      '/Users/Starry/Downloads/풀무원_디자인밀_AI챗봇_인텐트정의_보완_20260602.xlsx')
OUT_ROOT = os.path.expanduser(os.environ.get('CHATBOT_OUT', '~/Downloads/SB_챗봇'))
CIRC = '①②③④⑤⑥⑦⑧⑨⑩'

def build_index(target, outdir):
    """08(모듈순/명/유형)+09(컴포넌트명)+10(바인딩 컴포넌트) → 라벨 iframe 인덱스."""
    wb = openpyxl.load_workbook(XLSX)
    ws8, ws9, ws10 = wb['8_시나리오모듈'], wb['09_컴포넌트스키마'], wb['10_모듈콘텐츠바인딩']
    cname = {}
    for r in range(3, ws9.max_row + 1):
        c = ws9.cell(r, 1).value
        if c:
            cname[c] = ws9.cell(r, 2).value
    binds = defaultdict(list)
    for r in range(3, ws10.max_row + 1):
        code = ws10.cell(r, 1).value
        if code and str(code).startswith('IT'):
            binds[code].append(ws10.cell(r, 5).value)
    mods = []
    for r in range(3, ws8.max_row + 1):
        if ws8.cell(r, 2).value == target:
            mods.append(dict(code=ws8.cell(r, 1).value, step=ws8.cell(r, 4).value,
                             name=ws8.cell(r, 5).value, atype=ws8.cell(r, 7).value))
    mods.sort(key=lambda m: str(m['step']))
    rows = ''
    for i, m in enumerate(mods):
        comps = []
        for c in binds[m['code']]:
            if c and c not in comps:
                comps.append(c)
        clabel = ' + '.join(f'{c} {cname.get(c, "")}' for c in comps)
        rows += (f'  <h2>{CIRC[i]} {m["code"]} — {m["name"]} ({m["atype"]} / {clabel})</h2>\n'
                 f'  <div class="frame-wrap"><iframe src="{m["code"]}.html"></iframe></div>\n')
    doc = ('<!DOCTYPE html><html lang="ko"><head><meta charset="UTF-8">\n'
           f'<title>IT{target:03d} SB</title>\n<style>\n'
           '  body{margin:0;background:#6d7882;font-family:sans-serif;}\n'
           '  h2{color:#fff;margin:18px 0 6px 16px;font-size:15px;font-weight:600;}\n'
           '  .frame-wrap{width:1280px;height:760px;overflow:hidden;margin:0 16px 24px;border:1px solid #33363d;background:#fff;}\n'
           '  iframe{width:2560px;height:1520px;transform:scale(.5);transform-origin:top left;border:0;display:block;}\n'
           '</style></head>\n<body>\n' + rows + '</body></html>\n')
    with open(os.path.join(outdir, 'index.html'), 'w', encoding='utf-8') as f:
        f.write(doc)
    return len(mods)

def main():
    if len(sys.argv) < 2:
        sys.exit('사용법: python3 build_chatbot_sb.py <인텐트번호>   예) 7')
    target = int(sys.argv[1])
    tag = f'IT{target:03d}'
    out = os.path.join(OUT_ROOT, tag)
    mirror = f'/tmp/sb_view/{tag}'
    os.makedirs(out, exist_ok=True)
    json_path = os.path.join(HERE, '_build', f'{tag}_screens.json')

    # ① 변환
    print(f'① 변환  {tag} …')
    subprocess.run([sys.executable, os.path.join(HERE, 'chatbot_to_sb.py'), str(target)], check=True)
    # ② SB HTML
    print(f'② 생성  → {out}')
    subprocess.run([sys.executable, GEN, '--variant', 'mobile',
                    '--input', json_path, '--output', out], check=True)
    # ③ 인덱스
    n = build_index(target, out)
    print(f'③ 인덱스 index.html ({n} 화면)')
    # 미러(미리보기 서버용)
    if os.path.isdir(mirror):
        shutil.rmtree(mirror)
    shutil.copytree(out, mirror)

    print('\n완료')
    print(f'  · Finder/브라우저: {out}/index.html')
    print(f'  · 미리보기 서버:  http://127.0.0.1:8765/{tag}/index.html')

if __name__ == '__main__':
    main()
