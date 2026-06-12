#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""챗봇 컴포넌트 가이드 빌더 — 워크북(07 메타 + 09 계약 + 08/10 사용처) → 단일 HTML 카탈로그.

A안 레이아웃: 분류(텍스트/선택/카드/입력/배너/액션)별 그룹 + 좌측 고정 네비.
컴포넌트 카드 = [라이브 프리뷰 | 스펙 표]. 프리뷰는 chatbot_components.ARCHETYPES
렌더러를 그대로 호출하므로 실 SB와 100% 동일(드리프트 0).

데이터 주도: 시트 07의 컴포넌트 행 + 09의 렌더 계약 + ARCHETYPES 딕셔너리에서만 읽는다.
나중에 컴포넌트(07/09 행) + 렌더러(chatbot_components)만 추가하면 가이드에 자동 반영
— 이 스크립트는 수정 불필요.

  python3 build_component_guide.py                # 워크북 전체 인텐트에서 사용처 집계
  python3 build_component_guide.py --intents 7,9,10,14   # 특정 인텐트로 한정

출력: ~/Downloads/SB_챗봇/_GUIDE/component-guide.html  (CHATBOT_OUT 루트 하위)
"""
import os, sys, argparse, html
from collections import defaultdict, OrderedDict
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
import chatbot_components as cc
import generate_sb as gsb     # SB 양식 스캐폴드(메타 헤더+푸터+스타일) 재사용

XLSX = os.environ.get('CHATBOT_XLSX',
                      '/Users/Starry/Downloads/풀무원_디자인밀_AI챗봇_인텐트정의_보완_20260602.xlsx')
OUT_ROOT = os.path.expanduser(os.environ.get('CHATBOT_OUT', '~/Downloads/SB_챗봇'))

# 분류 표시 순서(없는 분류는 뒤에 등장순으로 append)
CAT_ORDER = ['텍스트', '선택', '카드', '입력', '배너', '액션']
CATEGORY_EN = {'텍스트': 'Text', '선택': 'Button', '카드': 'Card',
               '입력': 'Input', '배너': 'Banner', '액션': 'Action'}
SRC_TAG = {'static': '고정', 'api': 'API', 'session': '세션', 'rag': 'RAG', 'system': '시스템'}

# 분기(케이스) 정의 — 화면을 나누지 않고 컴포넌트 옆에 나란히 병치.
# {컴포넌트코드: [(케이스라벨, roles override), ...]}
CASES = {
    'U-11': [   # 결과배너: 성공 / 실패
        ('성공', {'status': 'success', 'message': '요청이 처리되었어요',
                 'detail': '변경 내용이 정상 반영되었습니다.'}),
        ('실패', {'status': 'fail', 'message': '처리하지 못했어요',
                 'detail': '잠시 후 다시 시도해주세요.'}),
    ],
    'U-10': [   # 로딩: 기본 / 취소 가능
        ('기본', {'loadingText': '잠시만 기다려 주세요', 'cancelable': 'false'}),
        ('취소 가능', {'loadingText': '주문 정보를 불러오고 있어요', 'cancelable': 'true'}),
    ],
}

# 프리뷰용 더미 샘플(chatbot_to_sb와 동일 — 실값은 런타임 바인딩) -----------------
import re
FIELD_SAMPLE = {
    'balance': '1,250 P', 'expiring': '80 P', 'scheduled': '300 P',
    'point': '1,250 P', 'amount': '38,000원',
    'result': 'success', 'status': 'success',
    'maxQuantity': '5', 'currentQty': '2', 'newQty': '3', 'quantity': '2',
    'selectedProduct': '유기농 콩나물 1+1', 'productName': '유기농 콩나물 1+1',
    'name': '유기농 콩나물 1+1', 'title': '유기농 콩나물 1+1',
    'deliveryDate': '6/10(화)', 'orderDate': '6/1(일)', 'date': '6/10(화)',
    'orderId': 'DM240601-0001', 'price': '12,900원',
}
def sample(v):
    s = (v or '').strip()
    def repl(m):
        key = m.group(1).split('.')[-1].rstrip('[]').strip()
        return FIELD_SAMPLE.get(key, '○○')
    return re.sub(r'\{\{\s*([^}]+?)\s*\}\}', repl, s)
H = cc.Helpers(sample)


def esc(v):
    return html.escape('' if v is None else str(v))


def load_workbook_sheets():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    return wb['07_UI컴포넌트'], wb['09_컴포넌트스키마'], wb['8_시나리오모듈'], wb['10_모듈콘텐츠바인딩']


def read_components(ws07):
    """07 → [{code,name,cat,defn,answer_types}] (등장 순서 보존)."""
    comps = []
    for r in range(3, ws07.max_row + 1):
        code = ws07.cell(r, 1).value
        if not code or not str(code).startswith('U-'):
            continue
        comps.append(dict(
            code=str(code).strip(), name=ws07.cell(r, 2).value, cat=ws07.cell(r, 3).value,
            defn=ws07.cell(r, 4).value, answer_types=ws07.cell(r, 6).value))
    return comps


def read_schema(ws09):
    """09 → archetype[comp], slots[comp]=[{slot,type,req,src,role,default,desc}], role[(comp,slot)]."""
    archetype, slots, role_of = {}, defaultdict(list), {}
    for r in range(3, ws09.max_row + 1):
        comp = ws09.cell(r, 1).value
        if not comp:
            continue
        comp = str(comp).strip()
        arch = ws09.cell(r, 10).value
        if arch:
            archetype[comp] = str(arch).strip()
        slot, role = ws09.cell(r, 3).value, ws09.cell(r, 11).value
        if slot:
            slots[comp].append(dict(
                slot=slot, type=ws09.cell(r, 4).value, req=ws09.cell(r, 5).value,
                default=ws09.cell(r, 6).value, src=ws09.cell(r, 7).value,
                desc=ws09.cell(r, 8).value, role=role))
            if role:
                role_of[(comp, slot)] = role
    return archetype, slots, role_of


def read_usage(ws08, ws10, intents):
    """08+10 → usage[comp]=[(module_code, intent_name)], bindings[(module_code,comp)]=[(slot,val,src)].
    intents=None이면 전체."""
    # 08: module_code -> (intent#, intent_name, step, name)
    meta = {}
    for r in range(3, ws08.max_row + 1):
        code = ws08.cell(r, 1).value
        if not code or not str(code).startswith('IT'):
            continue
        meta[str(code).strip()] = dict(
            intent=ws08.cell(r, 2).value, intent_name=ws08.cell(r, 3).value,
            step=ws08.cell(r, 4).value, name=ws08.cell(r, 5).value)
    usage = defaultdict(list)        # comp -> [(module_code, intent_name)]
    bindings = defaultdict(list)     # (module_code, comp) -> [(slot, value, src)]
    seen = defaultdict(set)
    for r in range(3, ws10.max_row + 1):
        code = ws10.cell(r, 1).value
        if not code or not str(code).startswith('IT'):
            continue
        code = str(code).strip()
        intent_no = ws10.cell(r, 2).value
        if intents is not None and intent_no not in intents:
            continue
        comp = ws10.cell(r, 5).value
        if not comp:
            continue
        comp = str(comp).strip()
        bindings[(code, comp)].append((ws10.cell(r, 6).value, ws10.cell(r, 8).value, ws10.cell(r, 7).value))
        if code not in seen[comp]:
            seen[comp].add(code)
            iname = meta.get(code, {}).get('intent_name', '')
            usage[comp].append((code, iname))
    return usage, bindings, meta


def _literal_score(rows):
    """리터럴(템플릿 {{}} 아님·비어있지 않음) 값 개수 — 프리뷰 가독성 점수."""
    return sum(1 for (_slot, v, _src) in rows
               if v not in (None, '') and '{{' not in str(v))

def preview_roles(comp, slots, role_of, usage, bindings):
    """프리뷰용 roles 결정: 실 바인딩 1순위(리터럴 값이 가장 많은 모듈 — 가독성),
    없으면 09 기본값으로 합성."""
    use = usage.get(comp)
    if use:
        # 사용 모듈 중 리터럴 값이 가장 많은 것을 대표 프리뷰로 선택
        best = max((c for c, _ in use),
                   key=lambda mc: _literal_score(bindings.get((mc, comp), [])),
                   default=None)
        if best:
            roles = {}
            for (slot, value, _src) in bindings.get((best, comp), []):
                roles[role_of.get((comp, slot), slot)] = value
            if roles:
                return roles, ('binding', best)
    # 합성: 09 기본값 → role. 렌더러 자체 폴백이 있어 빈 dict여도 시각화는 됨.
    roles = {}
    for s in slots.get(comp, []):
        if s.get('default') not in (None, ''):
            roles[s['role'] or s['slot']] = s['default']
    return roles, ('synth', None)


# ===== HTML 조각 =====
def cat_chip(cat):
    return f'<span class="chip chip-cat">{esc(cat)}</span>'

def arch_chip(arch, ok):
    cls = 'chip chip-arch' if ok else 'chip chip-todo'
    label = esc(arch) if arch else '—'
    return f'<span class="{cls}">{label}{"" if ok else " · 준비 중"}</span>'

def usage_chip(n):
    if n == 0:
        return '<span class="chip chip-unused">미사용</span>'
    return f'<span class="chip chip-used">사용 {n}곳</span>'

def slot_table(comp, slots):
    rows = ''
    for s in slots.get(comp, []):
        req = '●' if str(s.get('req') or '').strip() in ('Y', 'y', 'TRUE', 'True', '필수', '●') else '○'
        src = SRC_TAG.get(str(s.get('src') or '').strip().lower(), s.get('src') or '-')
        rows += (
            '<tr>'
            f'<td class="mono">{esc(s.get("slot"))}</td>'
            f'<td class="mono dim">{esc(s.get("type"))}</td>'
            f'<td class="ctr">{req}</td>'
            f'<td>{esc(src)}</td>'
            f'<td class="mono">{esc(s.get("role"))}</td>'
            '</tr>')
    if not rows:
        rows = '<tr><td colspan="5" class="dim">정의된 슬롯 없음</td></tr>'
    return (
        '<table class="spec-table"><thead><tr>'
        '<th>슬롯(prop)</th><th>타입</th><th>필수</th><th>소스</th><th>역할(role)</th>'
        '</tr></thead><tbody>' + rows + '</tbody></table>')

def usage_block(comp, usage):
    use = usage.get(comp, [])
    if not use:
        return '<div class="meta-line"><span class="meta-key">사용처</span><span class="dim">아직 바인딩된 인텐트 없음</span></div>'
    seen, items = set(), []
    for code, iname in use:
        if code in seen:
            continue
        seen.add(code)
        items.append(f'<span class="use-pill"><b>{esc(code)}</b> {esc(iname)}</span>')
    return f'<div class="meta-line"><span class="meta-key">사용처</span><div class="use-wrap">{"".join(items)}</div></div>'

def answer_types_block(types):
    if not types:
        return ''
    chips = ''.join(f'<span class="at-pill">{esc(t.strip())}</span>'
                    for t in str(types).replace('，', ',').split(',') if t.strip())
    return f'<div class="meta-line"><span class="meta-key">답변유형</span><div class="use-wrap">{chips}</div></div>'

# 단독 출력 금지 — 반드시 선행 텍스트 버블과 함께 제공해야 하는 아키타입.
# (선택버튼은 워크북상 항상 U-01/U-08 동반. 입력 위젯은 UX 규칙상 질문 버블 동반)
PROMPT_PAIRED = {'buttons', 'input-date', 'input-stepper', 'input-text'}

# 대표 모듈에 U-01 텍스트가 없을 때 쓰는 아키타입별 기본 안내 문구.
DEFAULT_PROMPT = {
    'buttons': '원하시는 항목을 선택해주세요',
    'input-date': '날짜를 선택해주세요',
    'input-stepper': '수량을 입력해주세요',
    'input-text': '내용을 입력해주세요',
}

def leading_prompt(rep_mod, bindings, role_of, archetype):
    """선행 텍스트 버블 문구 — 대표 모듈의 U-01(텍스트) 바인딩값 우선, 없으면 아키타입별 기본 안내."""
    if rep_mod:
        for (slot, value, _src) in bindings.get((rep_mod, 'U-01'), []):
            if role_of.get(('U-01', slot), slot) == 'primaryText' or slot == 'text':
                if value:
                    return sample(value)
    return DEFAULT_PROMPT.get(archetype, '원하시는 항목을 선택해주세요')

def preview_html(comp, archetype, roles, lead_text=None):
    renderer = cc.ARCHETYPES.get(archetype)
    if not renderer:
        return ('<div class="preview-todo">렌더러 준비 중<br>'
                '<span class="dim">chatbot_components에 아키타입 추가 시 자동 렌더</span></div>')
    try:
        inner = renderer(roles, H)
    except Exception as e:
        return f'<div class="preview-todo">프리뷰 오류: {esc(e)}</div>'
    rows = ''
    if lead_text:   # 동반 텍스트 버블(규칙: 선택버튼 등은 단독 출력 금지)
        rows += cc.bot_row(cc.bubble(esc(lead_text)))
    rows += cc.bot_row(inner)
    return f'<div class="preview-stage">{rows}</div>'

def component_card(c, archetype, slots, role_of, usage, bindings):
    comp = c['code']
    arch = archetype.get(comp)
    ok = bool(cc.ARCHETYPES.get(arch))
    roles, (src, rep_mod) = preview_roles(comp, slots, role_of, usage, bindings)
    n_use = len({code for code, _ in usage.get(comp, [])})
    lead_text = leading_prompt(rep_mod, bindings, role_of, arch) if arch in PROMPT_PAIRED else None
    note = '실 바인딩 프리뷰' if src == 'binding' else '샘플 프리뷰(기본값)'
    if lead_text:
        note += ' · 텍스트 버블 동반'
    return (
        f'<article class="cmp-card" id="{esc(comp)}">'
        '<div class="cmp-head">'
        f'<span class="cmp-code">{esc(comp)}</span>'
        f'<span class="cmp-name">{esc(c["name"])}</span>'
        f'{arch_chip(arch, ok)}{cat_chip(c["cat"])}{usage_chip(n_use)}'
        '</div>'
        f'<p class="cmp-defn">{esc(c["defn"])}</p>'
        '<div class="cmp-body">'
        f'<div class="cmp-preview"><div class="preview-label">{note}</div>{preview_html(comp, arch, roles, lead_text)}</div>'
        f'<div class="cmp-spec">{slot_table(comp, slots)}{answer_types_block(c["answer_types"])}{usage_block(comp, usage)}</div>'
        '</div></article>')


CSS = """
:root{
  --krds-primary-5:#ecf2fe;--krds-primary-50:#256ef4;--krds-primary-60:#0b50d0;
  --krds-gray-0:#fff;--krds-gray-5:#f4f5f6;--krds-gray-10:#e6e8ea;--krds-gray-20:#cdd1d5;
  --krds-gray-30:#b1b8be;--krds-gray-40:#8a949e;--krds-gray-50:#6d7882;--krds-gray-60:#58616a;
  --krds-gray-70:#464c53;--krds-gray-80:#33363d;--krds-gray-90:#1e2124;
  --krds-danger-5:#fdefec;--krds-danger-50:#de3412;--krds-success-5:#eaf6ec;--krds-success-50:#228738;
  --krds-radius-sm:4px;--krds-radius-md:6px;--krds-radius-lg:8px;
  --etribe-orange:#f37320;
  --krds-font:Pretendard,'Pretendard GOV','Pretendard Variable',-apple-system,BlinkMacSystemFont,system-ui,Roboto,sans-serif;
}
*,*::before,*::after{margin:0;padding:0;box-sizing:border-box;font-family:var(--krds-font);}
body{background:var(--krds-gray-5);color:var(--krds-gray-90);}
button{font-family:var(--krds-font);}
/* 헤더 */
.guide-header{background:var(--krds-gray-0);border-bottom:3px solid var(--etribe-orange);padding:28px 40px;}
.guide-header h1{font-size:24px;font-weight:800;color:var(--krds-gray-90);}
.guide-header .sub{margin-top:6px;font-size:14px;color:var(--krds-gray-50);}
.guide-stats{margin-top:16px;display:flex;gap:10px;flex-wrap:wrap;}
.stat{background:var(--krds-gray-5);border:1px solid var(--krds-gray-10);border-radius:var(--krds-radius-md);padding:8px 14px;font-size:13px;color:var(--krds-gray-70);}
.stat b{font-size:16px;color:var(--krds-primary-60);font-weight:800;margin-right:4px;}
/* 레이아웃 */
.guide-wrap{display:grid;grid-template-columns:200px 1fr;align-items:start;gap:0;max-width:1320px;margin:0 auto;}
.guide-nav{position:sticky;top:0;align-self:start;padding:28px 20px;height:100vh;overflow:auto;border-right:1px solid var(--krds-gray-10);background:var(--krds-gray-0);}
.guide-nav .nav-title{font-size:11px;font-weight:700;color:var(--krds-gray-40);letter-spacing:.08em;margin-bottom:10px;}
.guide-nav a{display:flex;justify-content:space-between;align-items:center;text-decoration:none;color:var(--krds-gray-70);font-size:14px;font-weight:600;padding:9px 12px;border-radius:var(--krds-radius-md);}
.guide-nav a:hover{background:var(--krds-primary-5);color:var(--krds-primary-60);}
.guide-nav a .cnt{font-size:12px;color:var(--krds-gray-40);font-weight:600;}
.guide-main{padding:28px 40px 80px;min-width:0;}
/* 분류 섹션 */
.cat-section{margin-bottom:40px;}
.cat-section > h2{font-size:18px;font-weight:800;color:var(--krds-gray-90);padding-bottom:10px;border-bottom:2px solid var(--krds-gray-90);margin-bottom:20px;display:flex;align-items:baseline;gap:10px;}
.cat-section > h2 .cat-cnt{font-size:13px;font-weight:600;color:var(--krds-gray-40);}
/* 컴포넌트 카드 */
.cmp-card{background:var(--krds-gray-0);border:1px solid var(--krds-gray-10);border-radius:var(--krds-radius-lg);padding:20px 22px;margin-bottom:18px;scroll-margin-top:20px;}
.cmp-head{display:flex;align-items:center;gap:10px;flex-wrap:wrap;}
.cmp-code{font-size:13px;font-weight:800;color:var(--krds-gray-0);background:var(--krds-gray-80);border-radius:var(--krds-radius-sm);padding:3px 9px;letter-spacing:.02em;}
.cmp-name{font-size:17px;font-weight:800;color:var(--krds-gray-90);}
.cmp-defn{margin:10px 0 16px;font-size:14px;color:var(--krds-gray-60);line-height:1.6;}
.chip{font-size:11px;font-weight:700;padding:3px 9px;border-radius:11px;line-height:1.4;}
.chip-cat{background:var(--krds-gray-5);color:var(--krds-gray-60);border:1px solid var(--krds-gray-20);}
.chip-arch{background:var(--krds-primary-5);color:var(--krds-primary-60);border:1px solid var(--krds-primary-50);}
.chip-todo{background:var(--krds-gray-5);color:var(--krds-gray-40);border:1px dashed var(--krds-gray-30);}
.chip-used{background:var(--krds-success-5);color:var(--krds-success-50);border:1px solid var(--krds-success-50);}
.chip-unused{background:var(--krds-gray-5);color:var(--krds-gray-40);border:1px solid var(--krds-gray-20);}
/* 카드 본문 2단 */
.cmp-body{display:grid;grid-template-columns:360px 1fr;gap:24px;align-items:start;}
.cmp-preview{border:1px solid var(--krds-gray-10);border-radius:var(--krds-radius-md);overflow:hidden;}
.preview-label{font-size:11px;font-weight:700;color:var(--krds-gray-40);background:var(--krds-gray-5);padding:6px 12px;border-bottom:1px solid var(--krds-gray-10);}
.preview-stage{background:var(--krds-gray-5);padding:18px 16px;display:flex;flex-direction:column;gap:12px;}
.preview-todo{background:var(--krds-gray-5);padding:32px 16px;text-align:center;font-size:13px;font-weight:700;color:var(--krds-gray-50);line-height:1.6;}
/* 스펙 */
.spec-table{width:100%;border-collapse:collapse;font-size:13px;}
.spec-table th{background:var(--krds-gray-5);color:var(--krds-gray-60);font-weight:700;text-align:left;padding:7px 10px;border:1px solid var(--krds-gray-10);white-space:nowrap;}
.spec-table td{padding:7px 10px;border:1px solid var(--krds-gray-10);color:var(--krds-gray-80);vertical-align:top;}
.spec-table td.ctr{text-align:center;}
.spec-table .mono{font-family:'SF Mono',Consolas,monospace;font-size:12px;}
.spec-table .dim{color:var(--krds-gray-40);}
.meta-line{display:flex;gap:10px;align-items:baseline;margin-top:12px;font-size:13px;}
.meta-key{flex-shrink:0;font-weight:700;color:var(--krds-gray-50);width:64px;}
.use-wrap{display:flex;gap:6px;flex-wrap:wrap;}
.use-pill{font-size:12px;background:var(--krds-gray-5);border:1px solid var(--krds-gray-10);border-radius:var(--krds-radius-sm);padding:3px 8px;color:var(--krds-gray-70);}
.use-pill b{color:var(--krds-gray-90);font-weight:800;}
.at-pill{font-size:12px;background:var(--krds-primary-5);border:1px solid var(--krds-primary-10);border-radius:var(--krds-radius-sm);padding:3px 8px;color:var(--krds-primary-60);font-weight:600;}
.dim{color:var(--krds-gray-40);}
@media(max-width:980px){.cmp-body{grid-template-columns:1fr;}.guide-wrap{grid-template-columns:1fr;}.guide-nav{display:none;}}
"""

FONT_LINKS = (
    '<link rel="stylesheet" as="style" crossorigin '
    'href="https://cdn.jsdelivr.net/gh/orioncactus/pretendard@v1.3.9/dist/web/static/pretendard-dynamic-subset.min.css">')


def group_by_cat(comps):
    """분류별 그룹 (CAT_ORDER 우선, 나머지 등장순)."""
    by_cat = OrderedDict((k, []) for k in CAT_ORDER)
    for c in comps:
        by_cat.setdefault((c['cat'] or '기타').strip(), []).append(c)
    return OrderedDict((k, v) for k, v in by_cat.items() if v)


def build(intents):
    ws07, ws09, ws08, ws10 = load_workbook_sheets()
    comps = read_components(ws07)
    archetype, slots, role_of = read_schema(ws09)
    usage, bindings, meta = read_usage(ws08, ws10, intents)

    by_cat = group_by_cat(comps)

    # 네비
    nav = '<div class="nav-title">분류</div>'
    for cat, items in by_cat.items():
        nav += f'<a href="#cat-{esc(cat)}">{esc(cat)}<span class="cnt">{len(items)}</span></a>'

    # 본문
    main = ''
    for cat, items in by_cat.items():
        cards = ''.join(component_card(c, archetype, slots, role_of, usage, bindings) for c in items)
        main += (f'<section class="cat-section" id="cat-{esc(cat)}">'
                 f'<h2>{esc(cat)} <span class="cat-cnt">{len(items)}개</span></h2>{cards}</section>')

    # 통계
    n_total = len(comps)
    n_ready = sum(1 for c in comps if cc.ARCHETYPES.get(archetype.get(c['code'])))
    n_used = sum(1 for c in comps if usage.get(c['code']))
    intent_label = '워크북 전체 인텐트' if intents is None else '인텐트 ' + ', '.join(str(i) for i in sorted(intents))

    stats = (
        f'<div class="stat"><b>{n_total}</b>컴포넌트</div>'
        f'<div class="stat"><b>{len(by_cat)}</b>분류</div>'
        f'<div class="stat"><b>{n_ready}</b>렌더 가능</div>'
        f'<div class="stat"><b>{n_used}</b>사용 중</div>')

    doc = (
        '<!DOCTYPE html><html lang="ko"><head><meta charset="UTF-8">'
        '<meta name="viewport" content="width=device-width, initial-scale=1.0">'
        '<title>디자인밀 AI 챗봇 — 컴포넌트 가이드</title>'
        f'{FONT_LINKS}<style>{CSS}</style></head><body>'
        '<header class="guide-header">'
        '<h1>디자인밀 AI 챗봇 — 컴포넌트 가이드</h1>'
        f'<div class="sub">KRDS 기반 챗봇 응답 컴포넌트 카탈로그 · 집계 범위: {esc(intent_label)}</div>'
        f'<div class="guide-stats">{stats}</div>'
        '</header>'
        '<div class="guide-wrap">'
        f'<nav class="guide-nav">{nav}</nav>'
        f'<main class="guide-main">{main}</main>'
        '</div></body></html>')
    return doc


# ===== SB 양식(ETRIBE) 출력 — 모든 컴포넌트를 한 화면에 취합 =====
def _render_inner(arch, roles, lead=None):
    """아키타입 렌더러 호출 → (선행 텍스트 버블) + 컴포넌트 inner HTML."""
    renderer = cc.ARCHETYPES.get(arch)
    if not renderer:
        return '<div style="color:var(--krds-gray-40);font-size:13px;padding:8px;">렌더러 준비 중</div>'
    try:
        inner = renderer(roles, H)
    except Exception as e:
        return f'<div style="color:var(--krds-danger-50);font-size:12px;">오류: {esc(e)}</div>'
    if lead:   # 선택버튼·입력 위젯 등은 텍스트 말풍선 동반(단독 출력 금지)
        return (f'<div style="display:flex;flex-direction:column;gap:10px;">'
                f'{cc.bubble(esc(lead))}{inner}</div>')
    return inner

def _case_cell(label, inner_html):
    tag = (f'<div style="font-size:12px;color:var(--krds-primary-60);font-weight:700;'
           f'background:var(--krds-primary-5);border-radius:11px;padding:3px 10px;'
           f'align-self:flex-start;margin-bottom:8px;">{esc(label)}</div>') if label else ''
    return (f'<div style="display:flex;flex-direction:column;">{tag}'
            f'<div style="background:var(--krds-gray-5);border:1px solid var(--krds-gray-10);'
            f'border-radius:var(--krds-radius-lg);padding:18px;min-width:240px;">{inner_html}</div></div>')

def _component_block(c, archetype, slots, role_of, usage, bindings):
    comp, arch = c['code'], archetype.get(c['code'])
    if comp in CASES:   # 분기: 케이스 나란히 병치
        cells = ''.join(_case_cell(lbl, _render_inner(arch, roles)) for lbl, roles in CASES[comp])
    else:
        roles, (src, rep_mod) = preview_roles(comp, slots, role_of, usage, bindings)
        lead = leading_prompt(rep_mod, bindings, role_of, arch) if arch in PROMPT_PAIRED else None
        cells = _case_cell(None, _render_inner(arch, roles, lead))
    head = (f'<div style="font-size:14px;color:var(--krds-gray-50);font-weight:700;margin-bottom:10px;">'
            f'#{esc(c["name"])} <span style="color:var(--krds-gray-30);font-weight:600;">{esc(comp)}</span></div>')
    return (f'<div style="display:flex;flex-direction:column;margin:0 40px 32px 0;">{head}'
            f'<div style="display:flex;gap:18px;align-items:flex-start;flex-wrap:wrap;">{cells}</div></div>')

# desc-panel 헬퍼 (sb-style-block의 desc-* 클래스 사용)
def _db(parts):  return '<li><p class="desc-block">' + '<br>'.join(parts) + '</p></li>'
def _l1(t):      return f'<span class="lvl1">{esc(t)}</span>'
def _l2(t):      return f'<span class="lvl2">&nbsp;&nbsp;&nbsp;• {esc(t)}</span>'
def _sec(title, body):
    return (f'<div class="desc-section"><div class="desc-section-header">{esc(title)}</div>'
            f'<div class="desc-section-body">{body}</div></div>')

def _desc_panel(comps, by_cat, n_ready):
    info = _sec('화면 정보', '<ul class="desc-list">'
                + _db([_l1('챗봇 컴포넌트 가이드 (SB)')])
                + _db([_l2('디자인밀 AI 챗봇 · KRDS 응답 컴포넌트'),
                       _l2(f'총 {len(comps)}종 / {len(by_cat)}분류 · 렌더 가능 {n_ready}종'),
                       _l2('프리뷰는 실 SB 렌더러와 동일(드리프트 0)')])
                + '</ul>')
    flow = _sec('UI 흐름', '<ul class="desc-list">'
                + _db([_l1('1. 진입 · 처리 대기'),
                       _l2('로딩인디케이터(U-10)로 API 응답 대기 표시'),
                       _l2('취소 가능 여부로 분기 — 케이스 병치')])
                + _db([_l1('2. 안내 메시지'),
                       _l2('텍스트말풍선(U-01) — 모든 봇 답변의 기본 단위')])
                + _db([_l1('3. 조회 결과 표시'),
                       _l2('단건: 단건카드(U-03) / 요약카드(U-08)'),
                       _l2('다건: 카드리스트(U-04) / 추천카드(U-13, 가로 스크롤)')])
                + _db([_l1('4. 사용자 입력 · 선택'),
                       _l2('선택버튼(U-02) · 날짜(U-05) · 수량(U-06) · 텍스트(U-07) · 결제(U-09)'),
                       _l2('★ 선택버튼·입력 위젯은 텍스트말풍선과 함께 제공(단독 출력 금지)')])
                + _db([_l1('5. 처리 결과'),
                       _l2('결과배너(U-11) — 성공 / 실패 분기(케이스 병치)')])
                + _db([_l1('6. 다음 행동 · 이동 · 예외'),
                       _l2('액션버튼(U-12) · 이동링크(U-14) · 이관카드(U-15) · 에러배너(U-16)')])
                + '</ul>')
    # 분기/케이스
    case_items = ''
    cmap = {c['code']: c for c in comps}
    for code, cases in CASES.items():
        nm = cmap.get(code, {}).get('name', code)
        case_items += _db([_l1(f'{code} {nm}'),
                           _l2(' / '.join(lbl for lbl, _ in cases))])
    branch = _sec('분기 / 케이스', '<ul class="desc-list">'
                  + _db([_l1('분기는 화면을 분리하지 않고 해당 컴포넌트 옆에 케이스로 나란히 배치')])
                  + case_items + '</ul>')
    # 분류별 컴포넌트
    cat_lis = ''
    for cat, items in by_cat.items():
        sub = [_l1(f'{CATEGORY_EN.get(cat, cat)} ({cat})')]
        for c in items:
            sub.append(_l2(f'{c["code"]} {c["name"]} — {c["defn"]}'))
        cat_lis += _db(sub)
    cats = _sec('분류별 컴포넌트', f'<ul class="desc-list">{cat_lis}</ul>')
    return f'<div class="desc-panel">{info}{flow}{branch}{cats}</div>'

def build_sb(intents):
    ws07, ws09, ws08, ws10 = load_workbook_sheets()
    comps = read_components(ws07)
    archetype, slots, role_of = read_schema(ws09)
    usage, bindings, _meta = read_usage(ws08, ws10, intents)
    by_cat = group_by_cat(comps)
    n_ready = sum(1 for c in comps if cc.ARCHETYPES.get(archetype.get(c['code'])))

    # 와이어프레임 보드: 분류 섹션 → 컴포넌트 블록(필요 시 케이스 병치)
    board = ''
    for cat, items in by_cat.items():
        blocks = ''.join(_component_block(c, archetype, slots, role_of, usage, bindings) for c in items)
        head = (f'<div style="font-size:24px;font-weight:800;color:var(--krds-gray-90);'
                f'padding-bottom:14px;border-bottom:2px solid var(--krds-gray-90);margin-bottom:26px;">'
                f'{esc(CATEGORY_EN.get(cat, cat))} '
                f'<span style="font-size:16px;font-weight:600;color:var(--krds-gray-40);">{esc(cat)}</span></div>')
        board += (f'<div style="margin-bottom:48px;">{head}'
                  f'<div style="display:flex;flex-wrap:wrap;">{blocks}</div></div>')
    wf = (f'<div class="wf-panel" style="align-items:stretch;padding:48px 56px 100px;">'
          f'<div style="display:flex;flex-direction:column;">{board}</div></div>')
    body = wf + _desc_panel(comps, by_cat, n_ready)

    skill_dir = os.path.dirname(HERE)
    from pathlib import Path
    gsb.verify_logo(Path(skill_dir))
    style_block = gsb.load_template(Path(skill_dir), 'sb-style-block.html')
    logo = gsb.load_template(Path(skill_dir), 'sb-logo-data-url.txt').strip()
    return gsb.render_html(
        screen_id='GUIDE-COMP-01', title='챗봇 컴포넌트 가이드',
        path_str='AI 챗봇 > 컴포넌트 가이드',
        author=os.environ.get('CHATBOT_AUTHOR', 'AX Pilot'), ymd='2026-06',
        body=body, style_block=style_block, logo_data_url=logo)


def main():
    ap = argparse.ArgumentParser(description='챗봇 컴포넌트 가이드 빌더')
    ap.add_argument('--intents', default=None,
                    help='사용처 집계 인텐트 한정 (예: 7,9,10,14). 미지정 시 워크북 전체.')
    ap.add_argument('--format', choices=['web', 'sb', 'both'], default='both',
                    help='web=웹 카탈로그, sb=ETRIBE SB 양식, both=둘 다(기본)')
    ap.add_argument('--output', default=None, help='출력 경로(단일 포맷일 때만)')
    args = ap.parse_args()

    intents = None
    if args.intents:
        intents = {int(x) for x in args.intents.replace(' ', '').split(',') if x}

    fmts = {'web': ['web'], 'sb': ['sb'], 'both': ['web', 'sb']}[args.format]
    single = len(fmts) == 1
    targets = {
        'web': ('component-guide.html', build),
        'sb': ('component-guide-sb.html', build_sb),
    }

    gdir = os.path.join(OUT_ROOT, '_GUIDE')
    os.makedirs(gdir, exist_ok=True)
    print('컴포넌트 가이드 생성 완료')
    for key in fmts:
        fname, fn = targets[key]
        out = args.output if (args.output and single) else os.path.join(gdir, fname)
        with open(out, 'w', encoding='utf-8') as f:
            f.write(fn(intents))
        print(f'  · [{key}] {out}')


if __name__ == '__main__':
    main()
