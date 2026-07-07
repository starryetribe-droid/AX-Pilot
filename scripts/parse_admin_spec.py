#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""어드민 세부기능정의 워크북 파서 + 커버리지 검증기.

워크북(★IA매핑: 항목/세부항목/기능/세부기능/비고/진행여부/예시 화면/기능매핑)을
구조화 spec JSON + 작성 체크리스트로 변환하고, 작성된 PRD/SB가 워크북의
기능 라벨·규칙·예외를 전부 커버하는지 기계 대조한다.

사용:
  # ① 파싱: spec JSON + 체크리스트 생성
  ADMIN_SPEC_XLSX="/path/정의서.xlsx" python3 parse_admin_spec.py --id ADM-CHAT

  # ② 검증: PRD/SB 파일이 spec을 전부 커버하는지 대조 (여러 파일 가능)
  python3 parse_admin_spec.py --id ADM-CHAT --verify PRD.md SB1.html ...

출력: scripts/_build/{ID}_spec.json · {ID}_checklist.md
검증 결과: 누락 라벨/규칙 목록 + exit code (누락 있으면 1)
"""
import argparse, difflib, json, os, re, sys, unicodedata

HERE = os.path.dirname(os.path.abspath(__file__))
BUILD = os.path.join(HERE, '_build')

XLSX = os.environ.get('ADMIN_SPEC_XLSX',
                      os.path.expanduser('~/Downloads/디자인밀챗봇_어드민_세부기능정의_20260629.xlsx'))

NUM_RE = re.compile(r'^\s*(\d+(?:-\d+)*)\s*[.)]\s*(.+?)\s*$')          # "1-2. 라벨"
EXC_RE = re.compile(r'없을 경우|없는 경우|전일 시|그 외|실패|오류|에러|예외|불가|미노출|비활성|만료|권한 없')
BR_RE  = re.compile(r'[가-힣\)] 시[ ,]|경우')                            # 조건 분기 휴리스틱


def norm(s):
    s = unicodedata.normalize('NFKC', str(s or ''))
    return re.sub(r'\s+', '', s)                                        # 공백 전부 제거해 비교


def parse(xlsx, sheet=None):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]

    # 헤더 행/컬럼 탐지 ('항목'과 '기능'이 함께 있는 행)
    hdr_row, cols = None, {}
    for r in range(1, min(ws.max_row, 10) + 1):
        vals = {c: str(ws.cell(r, c).value or '').strip() for c in range(1, ws.max_column + 1)}
        if '항목' in vals.values() and '기능' in vals.values():
            hdr_row = r
            for c, v in vals.items():
                if v:
                    cols[v] = c
            break
    assert hdr_row, '헤더 행(항목/기능)을 찾지 못함'

    def cell(r, name):
        c = cols.get(name)
        return str(ws.cell(r, c).value or '').strip() if c else ''

    blocks, menu = [], ''
    for r in range(hdr_row + 1, ws.max_row + 1):
        sub = cell(r, '세부항목')
        if not sub and not cell(r, '기능'):
            continue
        menu = cell(r, '항목') or menu                                   # 병합 셀 forward-fill
        feats = []                                                       # 기능 아웃라인
        for ln in cell(r, '기능').splitlines():
            m = NUM_RE.match(ln)
            if m:
                feats.append(dict(num=m.group(1), label=m.group(2)))
            elif ln.strip().lstrip('-').strip() and feats:
                feats[-1].setdefault('extra', []).append(ln.strip())
        if not feats and cell(r, '기능').strip():                        # 번호 없는 산문형 기능도 라벨로
            feats.append(dict(num='1', label=cell(r, '기능').splitlines()[0].strip()))
        rules = []                                                       # 세부기능 규칙 라인
        cur = ''
        for ln in cell(r, '세부기능').splitlines():
            if not ln.strip():
                continue
            m = NUM_RE.match(ln)
            if m:
                cur = m.group(1)
                text = m.group(2)
            else:
                text = ln.strip().lstrip('-').strip()
            kind = ('exception' if EXC_RE.search(text)
                    else 'branch' if BR_RE.search(text)
                    else 'rule')
            rules.append(dict(num=cur, text=text, kind=kind))
        blocks.append(dict(no=cell(r, '') or len(blocks) + 1, menu=menu, sub=sub,
                           features=feats, rules=rules,
                           note=cell(r, '비고'), status=cell(r, '진행여부'),
                           mapping=cell(r, '기능매핑')))
    return dict(source=os.path.basename(xlsx), sheet=ws.title, blocks=blocks)


def write_outputs(spec, fid):
    os.makedirs(BUILD, exist_ok=True)
    jp = os.path.join(BUILD, f'{fid}_spec.json')
    json.dump(spec, open(jp, 'w', encoding='utf-8'), ensure_ascii=False, indent=2)

    lines = [f'# {fid} 작성 체크리스트 — {spec["source"]}', '',
             '> PRD §5(예외)·§6(UI)와 SB Description에 아래 항목이 **전부** 반영돼야 한다.',
             '> 저장 전 `parse_admin_spec.py --verify` 로 기계 대조할 것.', '']
    for b in spec['blocks']:
        lines.append(f'## [{b["menu"]}] {b["sub"]}')
        if b['features']:
            lines.append('### 기능(라벨 — 하드 매칭)')
            lines += [f'- [ ] {f["num"]}. {f["label"]}' for f in b['features']]
        exc = [r for r in b['rules'] if r['kind'] == 'exception']
        br = [r for r in b['rules'] if r['kind'] == 'branch']
        if exc:
            lines.append('### 예외 (PRD §5 필수)')
            lines += [f'- [ ] ({r["num"]}) {r["text"]}' for r in exc]
        if br:
            lines.append('### 조건 분기')
            lines += [f'- [ ] ({r["num"]}) {r["text"]}' for r in br]
        rest = [r for r in b['rules'] if r['kind'] == 'rule']
        if rest:
            lines.append('### 규칙 (Default·포맷·옵션)')
            lines += [f'- [ ] ({r["num"]}) {r["text"]}' for r in rest]
        lines.append('')
    cp = os.path.join(BUILD, f'{fid}_checklist.md')
    open(cp, 'w', encoding='utf-8').write('\n'.join(lines))
    return jp, cp


def verify(spec, targets):
    """PRD/SB 텍스트에 라벨(하드)·규칙(퍼지) 커버리지 대조."""
    corpus = ''
    for t in targets:
        corpus += open(t, encoding='utf-8').read()
    ncorpus = norm(re.sub(r'<[^>]+>', ' ', corpus))                      # 태그 제거 후 정규화
    corpus_lines = [norm(l) for l in re.split(r'[\n<>]', corpus) if l.strip()]

    missing_labels, weak_rules = [], []
    for b in spec['blocks']:
        for f in b['features']:
            if norm(f['label']) not in ncorpus:
                missing_labels.append(f'[{b["sub"]}] {f["num"]}. {f["label"]}')
        for r in b['rules']:
            key = norm(r['text'])
            if key in ncorpus:
                continue
            best = max((difflib.SequenceMatcher(None, key, cl).ratio()
                        for cl in corpus_lines if cl), default=0)
            if best < 0.55:
                weak_rules.append((r['kind'], f'[{b["sub"]}] ({r["num"]}) {r["text"]}', round(best, 2)))

    print(f'검증 대상: {", ".join(os.path.basename(t) for t in targets)}')
    if missing_labels:
        print(f'\n✗ 누락 라벨 {len(missing_labels)}건 (반드시 반영):')
        [print('   -', m) for m in missing_labels]
    else:
        print('✓ 기능 라벨 전수 반영')
    if weak_rules:
        print(f'\n△ 매칭 약한 규칙 {len(weak_rules)}건 (반영 여부 직접 확인):')
        [print(f'   - ({k}) {t}  [유사도 {s}]') for k, t, s in weak_rules]
    else:
        print('✓ 규칙·예외 전수 매칭')
    return 1 if missing_labels else 0


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--xlsx', default=XLSX)
    ap.add_argument('--sheet', default=None)
    ap.add_argument('--id', required=True, help='기능 ID (예: ADM-CHAT)')
    ap.add_argument('--verify', nargs='*', help='커버리지 검증할 PRD/SB 파일들')
    a = ap.parse_args()

    if not os.path.exists(a.xlsx):
        sys.exit(f'[ERROR] 워크북 없음: {a.xlsx} (ADMIN_SPEC_XLSX 로 지정)')
    spec = parse(a.xlsx, a.sheet)
    jp, cp = write_outputs(spec, a.id)

    if a.verify:
        sys.exit(verify(spec, a.verify))

    n_feat = sum(len(b['features']) for b in spec['blocks'])
    n_exc = sum(1 for b in spec['blocks'] for r in b['rules'] if r['kind'] == 'exception')
    n_br = sum(1 for b in spec['blocks'] for r in b['rules'] if r['kind'] == 'branch')
    print(f'파싱 완료: 블록 {len(spec["blocks"])} · 기능 라벨 {n_feat} · 예외 {n_exc} · 분기 {n_br}')
    for b in spec['blocks']:
        print(f'  [{b["menu"]}] {b["sub"]}: 기능 {len(b["features"])} · 규칙 {len(b["rules"])}')
    print('spec:', jp)
    print('checklist:', cp)


if __name__ == '__main__':
    main()
